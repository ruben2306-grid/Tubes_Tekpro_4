import streamlit as st
import numpy as np
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

st.set_page_config(page_title="Kalkulator Finansial", layout="centered")
st.title("Kalkulator Finansial Sederhana")

menu = st.sidebar.selectbox("Pilih Simulasi", ["Tabungan", "Pinjaman"])
input_mode = st.radio("Sumber Input", ["Input Manual", "Upload Excel"])

def valid_format(teks):
    return re.match(r'^(\d{1,3}(\.\d{3})*|\d+)(,\d+)?$', teks)

def rupiah_to_float(teks):
    return float(teks.replace(".", "").replace(",", "."))

def format_id(x):
    s = f"{x:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")

# ================= TABUNGAN =================
if menu == "Tabungan":
    st.subheader("Simulasi Tabungan")

    if input_mode == "Input Manual":
        saldo_txt = st.text_input("Saldo Awal (Rp)")
        setoran_txt = st.text_input("Setoran Bulanan (Rp)")
        bunga_txt = st.text_input("Bunga Tahunan (%)")
        bulan = st.slider("Lama Menabung (bulan)", 1, 120, 12)
    else:
        file = st.file_uploader(
            "Upload Excel Tabungan (kolom: saldo_awal, setoran_bulanan, bunga_tahunan, lama_bulan)",
            type=["xlsx"]
        )

    if st.button("Submit Tabungan"):
        if input_mode == "Input Manual":
            if "" in [saldo_txt, setoran_txt, bunga_txt]:
                st.warning("Lengkapi data.")
                st.stop()
            saldo_awal = rupiah_to_float(saldo_txt)
            setoran = rupiah_to_float(setoran_txt)
            bunga = rupiah_to_float(bunga_txt) / 100 / 12
        else:
            if file is None:
                st.warning("Upload file Excel.")
                st.stop()
            df_in = pd.read_excel(file)
            saldo_awal = df_in.loc[0, "saldo_awal"]
            setoran = df_in.loc[0, "setoran_bulanan"]
            bunga = df_in.loc[0, "bunga_tahunan"] / 100 / 12
            bulan = int(df_in.loc[0, "lama_bulan"])

        bulan_arr = np.arange(0, bulan + 1)
        saldo = np.zeros(len(bulan_arr))
        saldo[0] = saldo_awal

        for i in range(1, len(bulan_arr)):
            saldo[i] = saldo[i - 1] * (1 + bunga) + setoran

        df = pd.DataFrame({
            "No": np.arange(1, len(bulan_arr) + 1),
            "Bulan": bulan_arr,
            "Saldo (Rp)": saldo
        })

        st.dataframe(df.assign(
            **{"Saldo (Rp)": df["Saldo (Rp)"].apply(format_id)}
        ))
        st.line_chart(df.set_index("Bulan")["Saldo (Rp)"])

        st.success(f"Total saldo akhir tabungan: Rp {format_id(saldo[-1])}")

        buf = BytesIO()
        df.to_excel(buf, index=False, sheet_name="Tabungan")
        buf.seek(0)

        wb = load_workbook(buf)
        ws = wb["Tabungan"]

        chart = LineChart()
        chart.title = "Grafik Saldo Tabungan"
        chart.x_axis.title = "Bulan"
        chart.y_axis.title = "Saldo (Rp)"

        y = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
        x = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)

        chart.add_data(y, titles_from_data=True)
        chart.set_categories(x)
        ws.add_chart(chart, "F2")

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        st.download_button(
            "📥 Download Excel Tabungan + Grafik",
            out,
            "hasil_tabungan.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ================= PINJAMAN =================
if menu == "Pinjaman":
    st.subheader("Simulasi Pinjaman")

    if input_mode == "Input Manual":
        pinjaman_txt = st.text_input("Jumlah Pinjaman (Rp)")
        angsuran_txt = st.text_input("Angsuran Bulanan (Rp)")
        bunga_txt = st.text_input("Bunga Tahunan (%)")
        tenor = st.slider("Tenor (bulan)", 1, 120, 12)
    else:
        file = st.file_uploader(
            "Upload Excel Pinjaman (kolom: pinjaman, angsuran, bunga_tahunan, tenor)",
            type=["xlsx"]
        )

    if st.button("Submit Pinjaman"):
        if input_mode == "Input Manual":
            if "" in [pinjaman_txt, angsuran_txt, bunga_txt]:
                st.warning("Lengkapi data.")
                st.stop()
            pinjaman = rupiah_to_float(pinjaman_txt)
            angsuran = rupiah_to_float(angsuran_txt)
            bunga = rupiah_to_float(bunga_txt) / 100 / 12
        else:
            if file is None:
                st.warning("Upload file Excel.")
                st.stop()
            df_in = pd.read_excel(file)
            pinjaman = df_in.loc[0, "pinjaman"]
            angsuran = df_in.loc[0, "angsuran"]
            bunga = df_in.loc[0, "bunga_tahunan"] / 100 / 12
            tenor = int(df_in.loc[0, "tenor"])

        sisa = pinjaman
        data = []

        for b in range(1, tenor + 1):
            bunga_bln = sisa * bunga
            pokok = angsuran - bunga_bln
            if pokok <= 0:
                st.error("Angsuran terlalu kecil.")
                break
            sisa = max(sisa - pokok, 0)
            data.append([b, b, sisa])
            if sisa == 0:
                break

        if data:
            df = pd.DataFrame(data, columns=["No", "Bulan", "Sisa Utang (Rp)"])

            st.dataframe(df.assign(
                **{"Sisa Utang (Rp)": df["Sisa Utang (Rp)"].apply(format_id)}
            ))
            st.line_chart(df.set_index("Bulan")["Sisa Utang (Rp)"])

            total_bayar = angsuran * len(df)
            total_bunga = total_bayar - pinjaman

            st.success(f"Total pembayaran: Rp {format_id(total_bayar)}")
            st.info(f"Total bunga: Rp {format_id(total_bunga)}")
            st.success(f"Lunas dalam {len(df)} bulan")

            buf = BytesIO()
            df.to_excel(buf, index=False, sheet_name="Pinjaman")
            buf.seek(0)

            wb = load_workbook(buf)
            ws = wb["Pinjaman"]

            chart = LineChart()
            chart.title = "Grafik Sisa Utang"
            chart.x_axis.title = "Bulan"
            chart.y_axis.title = "Sisa Utang (Rp)"

            y = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
            x = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)

            chart.add_data(y, titles_from_data=True)
            chart.set_categories(x)
            ws.add_chart(chart, "F2")

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            st.download_button(
                "📥 Download Excel Pinjaman + Grafik",
                out,
                "hasil_pinjaman.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
